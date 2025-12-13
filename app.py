from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import json
import re
from werkzeug.utils import secure_filename
from database import (
    init_database, get_data_for_api, add_item, update_item, delete_items,
    add_category, delete_category, get_item_by_id, import_from_excel_data,
    backup_database, list_backups, restore_database, delete_backup, cleanup_old_backups,
    update_category_order, update_item_order
)

# 尝试导入reportlab用于PDF导出
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# 导入配置模块（简化版，不再需要API key）
try:
    from config import load_config
except ImportError:
    def load_config():
        return {}

# 全局抑制fonttools的警告
import warnings
import logging
warnings.filterwarnings('ignore', message='.*CFF.*')
warnings.filterwarnings('ignore', category=UserWarning, module='fontTools')
logging.getLogger('fontTools').setLevel(logging.ERROR)

app = Flask(__name__)

# 支持环境变量配置数据目录（用于云平台持久化存储）
DATA_DIR = os.getenv('DATA_DIR', '.')
app.config['UPLOAD_FOLDER'] = os.path.join(DATA_DIR, 'uploads')
app.config['EXPORT_FOLDER'] = os.path.join(DATA_DIR, 'exports')

# 确保目录存在（Railway部署时需要）
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 确保目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)

# 全局变量：缓存字体注册状态（避免每次PDF生成都重新注册）
_CHINESE_FONT_REGISTERED = False
_CHINESE_FONT_NAME = 'Helvetica'
_TEMP_FONT_FILE = None

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 支持环境变量配置Excel文件路径（用于导入和导出）
EXCEL_FILE = os.getenv('EXCEL_FILE', '红玺台复式装修预算表.xlsx')

# 初始化数据库
# 注意：database.py 会自动检测并使用持久存储（/mnt）如果可用
init_database()

# 如果存在Excel文件且数据库为空，自动导入Excel数据
def migrate_excel_to_db_if_needed():
    """如果数据库为空且存在Excel文件，自动导入"""
    from database import get_all_items, get_all_categories
    items = get_all_items()
    categories = get_all_categories()
    
    # 如果数据库为空且Excel文件存在，则导入
    if not items and not categories and os.path.exists(EXCEL_FILE):
        try:
            print("检测到Excel文件，正在导入到数据库...")
            excel_data = parse_excel()
            import_from_excel_data(excel_data)
            print("✅ Excel数据已成功导入到数据库")
        except Exception as e:
            print(f"⚠️ Excel导入失败: {e}")

# 启动时执行迁移
migrate_excel_to_db_if_needed()

# 定期自动备份（使用后台线程）
import threading
import time

def auto_backup_worker():
    """后台线程：定期自动备份数据库"""
    while True:
        try:
            # 每24小时备份一次
            time.sleep(24 * 60 * 60)
            
            # 创建自动备份
            backup_info = backup_database('auto_backup')
            print(f"✅ 自动备份已创建: {backup_info['filename']}")
            
            # 清理旧备份（保留最新的20个）
            deleted_count = cleanup_old_backups(keep_count=20)
            if deleted_count > 0:
                print(f"🗑️ 已清理 {deleted_count} 个旧备份")
        except Exception as e:
            print(f"⚠️ 自动备份失败: {e}")

# 启动自动备份线程
backup_thread = threading.Thread(target=auto_backup_worker, daemon=True)
backup_thread.start()
print("✅ 自动备份服务已启动（每24小时备份一次）")

def validate_excel_format(file_path):
    """验证Excel文件格式是否符合要求"""
    errors = []
    warnings = []
    
    try:
        df = pd.read_excel(file_path, engine='openpyxl', header=None)
        
        if df.empty:
            errors.append('Excel文件为空')
            return {'valid': False, 'errors': errors, 'warnings': warnings}
        
        # 检查是否有分类行
        has_category = False
        has_header = False
        category_count = 0
        
        for i in range(len(df)):
            row = df.iloc[i].astype(str).tolist()
            first_col = str(row[0]).strip() if pd.notna(row[0]) else ''
            
            # 检查分类行
            if any(first_col.startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']):
                has_category = True
                category_count += 1
            
            # 检查表头行
            if first_col == '序号' and '项目' in str(row[1]):
                has_header = True
        
        if not has_category:
            errors.append('未找到分类行（应以"一、"、"二、"等开头）')
        
        if not has_header:
            errors.append('未找到表头行（应包含"序号"和"项目"列）')
        
        if category_count == 0:
            warnings.append('未找到任何分类，建议至少有一个分类')
        
        # 检查必需的列
        if has_header:
            header_row = None
            for i in range(len(df)):
                row = df.iloc[i].astype(str).tolist()
                if str(row[0]).strip() == '序号':
                    header_row = i
                    break
            
            if header_row is not None:
                headers = df.iloc[header_row].astype(str).tolist()
                required_columns = ['序号', '项目']
                missing_columns = []
                
                for req_col in required_columns:
                    if req_col not in ' '.join(headers):
                        missing_columns.append(req_col)
                
                if missing_columns:
                    errors.append(f'缺少必需的列：{", ".join(missing_columns)}')
        
        # 检查是否有数据行
        has_data = False
        for i in range(len(df)):
            row = df.iloc[i].astype(str).tolist()
            first_col = str(row[0]).strip() if pd.notna(row[0]) else ''
            if first_col.isdigit():
                has_data = True
                break
        
        if not has_data:
            warnings.append('未找到任何数据行（序号为数字的行）')
        
        valid = len(errors) == 0
        
        return {
            'valid': valid,
            'errors': errors,
            'warnings': warnings,
            'category_count': category_count,
            'has_data': has_data
        }
        
    except Exception as e:
        return {
            'valid': False,
            'errors': [f'文件读取失败: {str(e)}'],
            'warnings': []
        }

def parse_excel():
    """解析Excel文件，返回结构化的数据"""
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl', header=None)
    
    # 表头行（第4行，索引3）
    header_row = 3
    headers = ['序号', '项目', '单位', '预算数量', '预算费用', '当前投入', '最终花费', '差价', '备注']
    
    categories = []
    items = []
    current_category = None
    item_id = 0  # 内部ID，用于追踪
    
    for i in range(len(df)):
        row = df.iloc[i].astype(str).tolist()
        first_col = str(row[0]).strip() if pd.notna(row[0]) else ''
        
        # 检查是否是分类行（去掉"一、"、"二、"等前缀）
        if any(first_col.startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']):
            # 去掉前缀，只保留分类名称
            category_name = first_col
            for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']:
                if category_name.startswith(prefix):
                    category_name = category_name[len(prefix):].strip()
                    break
            current_category = category_name
            if current_category not in categories:
                categories.append(current_category)
        # 检查是否是数据行（序号是数字）
        elif first_col.isdigit() and i > header_row:
            try:
                seq_num = int(float(first_col))
                # 读取原始列（兼容旧格式）
                # Excel列顺序：序号(0), 项目(1), 单位(2), 预算数量(3), 1st预算(4), 2nd预算(5), 最终实际花费(6), 差价(7), 备注(8)
                val_1st = str(row[4]).strip() if pd.notna(row[4]) and str(row[4]) != 'nan' else ''
                val_2nd = str(row[5]).strip() if len(row) > 5 and pd.notna(row[5]) and str(row[5]) != 'nan' else ''
                val_old_actual = str(row[6]).strip() if len(row) > 6 and pd.notna(row[6]) and str(row[6]) != 'nan' else ''  # 旧格式：最终实际花费
                val_diff = str(row[7]).strip() if len(row) > 7 and pd.notna(row[7]) and str(row[7]) != 'nan' else ''  # 差价
                val_remark = str(row[8]).strip() if len(row) > 8 and pd.notna(row[8]) and str(row[8]) != 'nan' else ''  # 备注
                
                # 合并1st和2nd预算为预算费用：优先使用2nd，否则使用1st
                if val_2nd and val_2nd.replace('.','').replace('-','').isdigit():
                    budget_value = val_2nd
                elif val_1st and val_1st.replace('.','').replace('-','').isdigit():
                    budget_value = val_1st
                else:
                    budget_value = ''
                
                # 当前投入：从旧格式的"最终实际花费"读取，但如果等于预算值，说明可能是错误数据，设为空
                current_value = val_old_actual
                if budget_value and current_value:
                    try:
                        budget_num = float(budget_value)
                        current_num = float(current_value)
                        # 如果当前投入等于预算，可能是错误数据，设为空（默认为0）
                        if abs(budget_num - current_num) < 0.01:
                            current_value = ''
                    except:
                        pass
                
                # 最终花费：旧格式Excel中没有此列，默认为空
                final_value = ''
                
                item = {
                    'id': item_id,
                    'row_index': i,  # 原始行索引
                    'category': current_category or '未分类',
                    '序号': seq_num,
                    '项目': str(row[1]).strip() if pd.notna(row[1]) and str(row[1]) != 'nan' else '',
                    '单位': str(row[2]).strip() if pd.notna(row[2]) and str(row[2]) != 'nan' else '',
                    '预算数量': str(row[3]).strip() if pd.notna(row[3]) and str(row[3]) != 'nan' else '',
                    '预算费用': budget_value,
                    '当前投入': current_value,
                    '最终花费': final_value,
                    '差价': val_diff,
                    '备注': val_remark
                }
                # 清理空值
                for key in item:
                    if item[key] == 'nan' or item[key] == '':
                        item[key] = ''
                items.append(item)
                item_id += 1
            except:
                pass
    
    return {
        'categories': categories,
        'items': items,
        'headers': headers
    }

def save_excel(data):
    """保存数据到Excel，保持原始格式"""
    # 读取原始Excel文件
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # 创建项目ID到行索引的映射
    item_map = {}
    for item in data['items']:
        if 'row_index' in item:
            item_map[item['id']] = item['row_index']
    
        # 更新数据行
        for item in data['items']:
            if 'row_index' in item and item['row_index'] in item_map.values():
                row_idx = item['row_index'] + 1  # openpyxl使用1-based索引
                ws.cell(row=row_idx, column=1, value=item['序号'])
                ws.cell(row=row_idx, column=2, value=item['项目'])
                ws.cell(row=row_idx, column=3, value=item['单位'] if item['单位'] else None)
                ws.cell(row=row_idx, column=4, value=item['预算数量'] if item['预算数量'] else None)
                
                # 列顺序：预算费用(5), 当前投入(6), 最终花费(7), 差价(8)
                # 兼容旧格式
                val_1st = float(item.get('1st预算费用', 0) or 0) if item.get('1st预算费用') and str(item.get('1st预算费用')).replace('.','').replace('-','').isdigit() else 0
                val_2nd = float(item.get('2nd预算费用', 0) or 0) if item.get('2nd预算费用') and str(item.get('2nd预算费用')).replace('.','').replace('-','').isdigit() else 0
                val_budget = float(item.get('预算费用', 0) or 0) if item.get('预算费用') and str(item.get('预算费用')).replace('.','').replace('-','').isdigit() else (val_2nd if val_2nd > 0 else val_1st)
                val_current = float(item.get('当前投入', 0) or 0) if item.get('当前投入') and str(item.get('当前投入')).replace('.','').replace('-','').isdigit() else (float(item.get('最终实际花费', 0) or 0) if item.get('最终实际花费') and str(item.get('最终实际花费')).replace('.','').replace('-','').isdigit() else 0)
                val_final = float(item.get('最终花费', 0) or 0) if item.get('最终花费') and str(item.get('最终花费')).replace('.','').replace('-','').isdigit() else 0
                
                # 自动计算差价（预算费用 - 最终花费）
                val_diff = val_budget - val_final
                
                ws.cell(row=row_idx, column=5, value=val_budget if val_budget > 0 else None)  # 预算费用
                ws.cell(row=row_idx, column=6, value=val_current if val_current > 0 else None)  # 当前投入
                ws.cell(row=row_idx, column=7, value=val_final if val_final > 0 else None)  # 最终花费
                ws.cell(row=row_idx, column=8, value=val_diff if val_diff != 0 else None)  # 差价（自动计算）
                ws.cell(row=row_idx, column=9, value=item['备注'] if item['备注'] else None)
    
    # 更新合计
    update_totals_in_excel()
    
    wb.save(EXCEL_FILE)
    wb.close()

def add_item_to_excel(item_data, category):
    """在Excel中添加新项目"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # 找到分类所在的行（支持带前缀和不带前缀的查找）
    category_row = None
    for i in range(1, ws.max_row + 1):
        first_cell = safe_get_cell_value(ws, i, 1)
        if first_cell:
            cell_value = str(first_cell).strip()
            # 去掉前缀后匹配
            category_name = cell_value
            for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']:
                if category_name.startswith(prefix):
                    category_name = category_name[len(prefix):].strip()
                    break
            # 匹配分类名称
            if category_name == category:
                category_row = i
                break
    
    if category_row:
        # 找到该分类下的最后一个数据行
        insert_row = category_row + 1
        for i in range(category_row + 1, ws.max_row + 1):
            first_cell = safe_get_cell_value(ws, i, 1)
            if first_cell and (str(first_cell).strip() == '合计' or 
                              any(str(first_cell).strip().startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、'])):
                insert_row = i
                break
            elif first_cell and str(first_cell).strip().isdigit():
                insert_row = i + 1
        
        # 获取该分类下的最大序号
        max_seq = 0
        for i in range(category_row + 1, insert_row):
            seq_val = safe_get_cell_value(ws, i, 1)
            if seq_val and str(seq_val).strip().isdigit():
                max_seq = max(max_seq, int(float(str(seq_val).strip())))
        
        # 插入新行
        ws.insert_rows(insert_row)
        ws.cell(insert_row, 1, value=max_seq + 1)
        ws.cell(insert_row, 2, value=item_data.get('项目', ''))
        ws.cell(insert_row, 3, value=item_data.get('单位', '') if item_data.get('单位') else None)
        ws.cell(insert_row, 4, value=item_data.get('预算数量', '') if item_data.get('预算数量') else None)
        
        # 列顺序：预算费用(5), 当前投入(6), 最终花费(7), 差价(8)
        val_budget = float(item_data['预算费用']) if item_data.get('预算费用') and str(item_data['预算费用']).replace('.','').replace('-','').isdigit() else 0
        val_current = float(item_data['当前投入']) if item_data.get('当前投入') and str(item_data['当前投入']).replace('.','').replace('-','').isdigit() else 0
        val_final = float(item_data['最终花费']) if item_data.get('最终花费') and str(item_data['最终花费']).replace('.','').replace('-','').isdigit() else 0
        
        # 自动计算差价（预算费用 - 最终花费）
        val_diff = val_budget - val_final
        
        ws.cell(insert_row, 5, value=val_budget if val_budget > 0 else None)  # 预算费用
        ws.cell(insert_row, 6, value=val_current if val_current > 0 else None)  # 当前投入
        ws.cell(insert_row, 7, value=val_final if val_final > 0 else None)  # 最终花费
        ws.cell(insert_row, 8, value=val_diff if val_diff != 0 else None)  # 差价（自动计算）
        ws.cell(insert_row, 9, value=item_data.get('备注', '') if item_data.get('备注') else None)
        
        # 更新合计
        update_totals_in_excel()
    
    wb.save(EXCEL_FILE)
    wb.close()

def add_category_to_excel(category_name):
    """在Excel中添加新分类"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # 找到最后一个分类行的位置和序号
    last_category_row = 0
    last_category_num = 0
    
    for i in range(1, ws.max_row + 1):
        first_cell = safe_get_cell_value(ws, i, 1)
        if first_cell:
            first_col = str(first_cell).strip()
            if any(first_col.startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']):
                last_category_row = i
                # 提取序号
                for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']:
                    if first_col.startswith(prefix):
                        num_map = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}
                        last_category_num = max(last_category_num, num_map.get(prefix[0], 0))
                        break
    
    # 找到最后一个分类的合计行位置
    insert_row = ws.max_row + 1  # 默认插入到最后
    if last_category_row > 0:
        # 从最后一个分类行开始查找合计行
        for i in range(last_category_row + 1, ws.max_row + 1):
            first_cell = safe_get_cell_value(ws, i, 1)
            if first_cell:
                first_col = str(first_cell).strip()
                if first_col == '合计':
                    # 找到合计行，在其后插入
                    insert_row = i + 1
                    break
                # 如果遇到下一个分类，停止查找
                if any(first_col.startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']):
                    # 在遇到下一个分类之前插入
                    insert_row = i
                    break
    
    # 生成新分类的序号前缀
    num_to_prefix = {1: '一、', 2: '二、', 3: '三、', 4: '四、', 5: '五、', 6: '六、', 7: '七、', 8: '八、', 9: '九、', 10: '十、'}
    new_prefix = num_to_prefix.get(last_category_num + 1, '')
    
    # 插入分类行
    ws.insert_rows(insert_row)
    ws.cell(insert_row, 1, value=f"{new_prefix}{category_name}")
    
    # 插入表头行
    ws.insert_rows(insert_row + 1)
    ws.cell(insert_row + 1, 1, value='序号')
    ws.cell(insert_row + 1, 2, value='项目')
    ws.cell(insert_row + 1, 3, value='单位')
    ws.cell(insert_row + 1, 4, value='预算数量')
    ws.cell(insert_row + 1, 5, value='预算费用')
    ws.cell(insert_row + 1, 6, value='当前投入')
    ws.cell(insert_row + 1, 7, value='最终花费')
    ws.cell(insert_row + 1, 8, value='差价')
    ws.cell(insert_row + 1, 9, value='备注：选购意向（网购/实体店，品牌，型号等）')
    
    # 插入合计行
    ws.insert_rows(insert_row + 2)
    ws.cell(insert_row + 2, 1, value='合计')
    
    wb.save(EXCEL_FILE)
    wb.close()

def delete_items_from_excel(row_indices):
    """从Excel中删除项目，严格保护分类行、表头行和合计行"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # 先扫描所有受保护的行（分类行、表头行、合计行）
    protected_row_indices = set()
    
    for i in range(1, ws.max_row + 1):
        first_cell = safe_get_cell_value(ws, i, 1)
        if first_cell:
            first_col = str(first_cell).strip()
            row_idx = i - 1  # 转换为0-based索引
            
            # 检查是否是分类行（一、二、三等开头）
            if any(first_col.startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']):
                protected_row_indices.add(row_idx)
                continue
            
            # 检查是否是表头行（包含"序号"）
            if first_col == '序号':
                protected_row_indices.add(row_idx)
                continue
            
            # 检查是否是合计行
            if first_col == '合计':
                protected_row_indices.add(row_idx)
                continue
    
    # 过滤掉受保护的行
    safe_row_indices = [idx for idx in row_indices if idx not in protected_row_indices]
    blocked_count = len(row_indices) - len(safe_row_indices)
    
    if not safe_row_indices:
        wb.close()
        raise ValueError(f'不能删除分类行、表头行或合计行（尝试删除 {blocked_count} 个受保护的行）')
    
    # 按倒序删除，避免索引变化
    for row_idx in sorted(safe_row_indices, reverse=True):
        ws.delete_rows(row_idx + 1)  # openpyxl使用1-based索引
    
    # 删除后重新计算合计
    update_totals_in_excel()
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    if blocked_count > 0:
        return f'已删除 {len(safe_row_indices)} 项，跳过了 {blocked_count} 个受保护的行（分类行/表头行/合计行）'
    return None

def safe_get_cell_value(ws, row, col):
    """安全地获取单元格值，处理合并单元格"""
    try:
        # 首先检查该单元格是否在合并范围内
        for merged_range in ws.merged_cells.ranges:
            if (row >= merged_range.min_row and row <= merged_range.max_row and 
                col >= merged_range.min_col and col <= merged_range.max_col):
                # 如果在合并范围内，返回合并区域左上角单元格的值
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        
        # 如果不在合并范围内，直接获取单元格值
        cell = ws.cell(row, col)
        # 检查是否是MergedCell对象（这种情况不应该发生，但为了安全）
        if hasattr(cell, 'value'):
            return cell.value
        else:
            # 如果是MergedCell对象，尝试获取合并区域的主单元格值
            for merged_range in ws.merged_cells.ranges:
                if (row >= merged_range.min_row and row <= merged_range.max_row and 
                    col >= merged_range.min_col and col <= merged_range.max_col):
                    return ws.cell(merged_range.min_row, merged_range.min_col).value
            return None
    except Exception:
        # 如果所有方法都失败，返回None
        return None

def safe_set_cell_value(ws, row, col, value):
    """安全地设置单元格值，处理合并单元格"""
    try:
        # 检查该单元格是否在合并范围内
        for merged_range in ws.merged_cells.ranges:
            if (row >= merged_range.min_row and row <= merged_range.max_row and 
                col >= merged_range.min_col and col <= merged_range.max_col):
                # 如果在合并范围内，只更新合并区域的主单元格（左上角）
                if row == merged_range.min_row and col == merged_range.min_col:
                    ws.cell(row, col, value=value)
                # 否则跳过（不更新合并单元格的非主单元格）
                return
        # 如果不在合并范围内，直接更新
        ws.cell(row, col, value=value)
    except Exception:
        # 如果更新失败，尝试直接更新（忽略错误）
        pass

def normalize_imported_data():
    """规范化导入的数据：设置默认值并同步2nd预算"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    header_row = 3  # 表头行（第4行，索引3）
    
    # 遍历所有数据行
    for i in range(header_row + 1, ws.max_row + 1):
        first_cell = safe_get_cell_value(ws, i, 1)
        # 检查是否是数据行（序号是数字）
        if first_cell and str(first_cell).strip().isdigit():
            # 列顺序：预算费用(5), 当前投入(6), 最终花费(7), 差价(8)
            # 兼容旧格式：如果列5和列6都有值，合并为预算费用
            val_1st = safe_get_cell_value(ws, i, 5)
            val_2nd = safe_get_cell_value(ws, i, 6)
            val_current = safe_get_cell_value(ws, i, 7)  # 可能是旧的实际花费或新的当前投入
            val_final = safe_get_cell_value(ws, i, 8) if ws.max_column >= 8 else None  # 最终花费可能在列8或列9
            
            # 检查值是否为空（None或空字符串）
            def is_empty(val):
                return val is None or (isinstance(val, str) and not val.strip())
            
            # 处理预算费用：合并1st和2nd，优先使用2nd
            val_budget = 0
            if not is_empty(val_2nd):
                try:
                    val_budget = float(val_2nd)
                except (ValueError, TypeError):
                    if not is_empty(val_1st):
                        try:
                            val_budget = float(val_1st)
                        except (ValueError, TypeError):
                            val_budget = 0
            elif not is_empty(val_1st):
                try:
                    val_budget = float(val_1st)
                except (ValueError, TypeError):
                    val_budget = 0
            
            # 处理当前投入：如果为空，设为0
            if is_empty(val_current):
                val_current = 0
            else:
                try:
                    val_current = float(val_current)
                except (ValueError, TypeError):
                    val_current = 0
            
            # 处理最终花费：如果为空，设为0
            if is_empty(val_final):
                val_final = 0
            else:
                try:
                    val_final = float(val_final)
                except (ValueError, TypeError):
                    val_final = 0
            
            # 计算差价（预算费用 - 最终花费）
            val_diff = val_budget - val_final
            
            # 更新Excel中的值（保留0值，因为0是有效的）
            safe_set_cell_value(ws, i, 5, val_budget if val_budget > 0 else None)  # 预算费用
            safe_set_cell_value(ws, i, 6, val_current if val_current > 0 else None)  # 当前投入
            safe_set_cell_value(ws, i, 7, val_final if val_final > 0 else None)  # 最终花费
            safe_set_cell_value(ws, i, 8, val_diff if val_diff != 0 else None)  # 差价
    
    # 更新合计
    update_totals_in_excel()
    
    wb.save(EXCEL_FILE)
    wb.close()

def update_totals_in_excel():
    """更新所有合计行的数值"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # 找到所有分类行和对应的合计行
    category_rows = []
    for i in range(1, ws.max_row + 1):
        first_cell = safe_get_cell_value(ws, i, 1)
        if first_cell:
            first_col = str(first_cell).strip()
            if any(first_col.startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']):
                category_rows.append(i)
    
    # 为每个分类计算合计
    for cat_row in category_rows:
        # 找到该分类下的合计行
        total_row = None
        for i in range(cat_row + 1, ws.max_row + 1):
            first_cell = safe_get_cell_value(ws, i, 1)
            if first_cell:
                first_col = str(first_cell).strip()
                if first_col == '合计':
                    total_row = i
                    break
                # 如果遇到下一个分类，停止查找
                if any(first_col.startswith(prefix) for prefix in ['一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '九、', '十、']):
                    break
        
        if total_row:
            # 计算该分类下所有项目的合计
            total_budget = 0
            total_current = 0
            total_final = 0
            total_diff = 0
            
            # 从分类行后到合计行前，累加所有数字序号行的费用
            for i in range(cat_row + 1, total_row):
                first_cell = safe_get_cell_value(ws, i, 1)
                if first_cell and str(first_cell).strip().isdigit():
                    # 列顺序：预算费用(5), 当前投入(6), 最终花费(7), 差价(8)
                    val_budget = safe_get_cell_value(ws, i, 5)
                    val_current = safe_get_cell_value(ws, i, 6)
                    val_final = safe_get_cell_value(ws, i, 7)
                    
                    if val_budget and isinstance(val_budget, (int, float)):
                        total_budget += float(val_budget)
                    if val_current and isinstance(val_current, (int, float)):
                        total_current += float(val_current)
                    if val_final and isinstance(val_final, (int, float)):
                        total_final += float(val_final)
            
            # 计算差价合计（预算费用 - 最终花费）
            total_diff = total_budget - total_final
            
            # 更新合计行（列顺序：预算费用(5), 当前投入(6), 最终花费(7), 差价(8)）
            safe_set_cell_value(ws, total_row, 5, total_budget if total_budget > 0 else None)  # 预算费用
            safe_set_cell_value(ws, total_row, 6, total_current if total_current > 0 else None)  # 当前投入
            safe_set_cell_value(ws, total_row, 7, total_final if total_final > 0 else None)  # 最终花费
            safe_set_cell_value(ws, total_row, 8, total_diff if total_diff != 0 else None)  # 差价
    
    wb.save(EXCEL_FILE)
    wb.close()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/verify-password', methods=['POST'])
def verify_password():
    """验证访问密码"""
    try:
        data = request.get_json()
        password = data.get('password', '')
        
        # 从环境变量获取密码，如果没有设置则使用默认密码
        correct_password = os.getenv('APP_PASSWORD', '902124')
        
        if password == correct_password:
            return jsonify({'success': True, 'message': '验证成功'})
        else:
            return jsonify({'success': False, 'error': '密码错误'}), 401
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/load', methods=['GET'])
def load_data():
    """加载数据库数据"""
    try:
        data = get_data_for_api()
        return jsonify({
            'success': True,
            'categories': data['categories'],
            'items': data['items'],
            'headers': data['headers'],
            'category_map': data.get('category_map', {})  # 添加分类ID映射
        })
    except Exception as e:
        import traceback
        error_msg = str(e)
        # 提供更友好的错误信息
        if 'disk i/o error' in error_msg.lower() or 'io error' in error_msg.lower():
            error_msg = '数据库I/O错误: 无法访问数据库文件。可能是存储挂载问题、权限问题或存储空间不足。请检查存储配置或联系管理员。'
        elif 'database is locked' in error_msg.lower():
            error_msg = '数据库被锁定: 可能有其他操作正在进行。请稍后重试。'
        return jsonify({'success': False, 'error': error_msg, 'traceback': traceback.format_exc()}), 500

@app.route('/api/add', methods=['POST'])
def add_item_route():
    """添加新项目"""
    try:
        data = request.json
        item = data.get('item', {})
        category = data.get('category', '')
        
        # 计算差价
        budget_cost = float(item.get('预算费用', 0) or 0)
        final_cost = float(item.get('最终花费', 0) or 0)
        item['差价'] = str(budget_cost - final_cost)
        
        # 调用数据库模块的add_item函数
        from database import add_item as db_add_item
        db_add_item(item, category)
        
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/add-category', methods=['POST'])
def add_category_route():
    """添加新分类"""
    try:
        data = request.json
        category_name = data.get('category_name', '').strip()
        
        if not category_name:
            return jsonify({'error': '分类名称不能为空'}), 400
        
        add_category(category_name)
        
        return jsonify({'success': True, 'message': '分类添加成功'})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/delete-category', methods=['POST'])
def delete_category_route():
    """删除分类"""
    try:
        data = request.json
        category_id = data.get('category_id')
        
        if category_id is None:
            return jsonify({'error': '请提供分类ID'}), 400
        
        try:
            category_id = int(category_id)
        except (ValueError, TypeError):
            return jsonify({'error': '无效的分类ID'}), 400
        
        result_message = delete_category(category_id)
        
        return jsonify({'success': True, 'message': result_message})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/backup', methods=['POST'])
def backup_route():
    """创建数据库备份"""
    try:
        data = request.json or {}
        description = data.get('description', '').strip()
        
        backup_info = backup_database(description)
        
        # 自动清理旧备份（保留最新的20个）
        cleanup_old_backups(keep_count=20)
        
        return jsonify({
            'success': True,
            'message': f'备份创建成功: {backup_info["filename"]}',
            'backup': backup_info
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/backups', methods=['GET'])
def list_backups_route():
    """列出所有备份"""
    try:
        backups = list_backups()
        return jsonify({'success': True, 'backups': backups})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/restore', methods=['POST'])
def restore_route():
    """恢复数据库"""
    try:
        data = request.json
        backup_filename = data.get('backup_filename')
        
        if not backup_filename:
            return jsonify({'error': '请提供备份文件名'}), 400
        
        result_message = restore_database(backup_filename)
        
        return jsonify({'success': True, 'message': result_message})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/delete-backup', methods=['POST'])
def delete_backup_route():
    """删除备份"""
    try:
        data = request.json
        backup_filename = data.get('backup_filename')
        
        if not backup_filename:
            return jsonify({'error': '请提供备份文件名'}), 400
        
        result_message = delete_backup(backup_filename)
        
        return jsonify({'success': True, 'message': result_message})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/update-category-order', methods=['POST'])
def update_category_order_route():
    """更新分类排序"""
    try:
        data = request.json
        category_orders = data.get('orders', [])
        
        if not category_orders:
            return jsonify({'error': '请提供分类排序数据'}), 400
        
        update_category_order(category_orders)
        return jsonify({'success': True, 'message': '分类排序已更新'})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/update-item-order', methods=['POST'])
def update_item_order_route():
    """更新项目排序"""
    try:
        data = request.json
        category_id = data.get('category_id')
        item_orders = data.get('orders', [])
        
        if category_id is None:
            return jsonify({'error': '请提供分类ID'}), 400
        if not item_orders:
            return jsonify({'error': '请提供项目排序数据'}), 400
        
        update_item_order(category_id, item_orders)
        return jsonify({'success': True, 'message': '项目排序已更新'})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/update', methods=['POST'])
def update_item_route():
    """更新项目"""
    try:
        data = request.json
        item = data.get('item', {})
        
        # 找到要更新的项目
        item_id = item.get('id')
        
        if not item_id:
            return jsonify({'error': '项目ID不能为空'}), 400
        
        # 检查项目是否存在
        existing_item = get_item_by_id(item_id)
        if not existing_item:
            return jsonify({'error': '项目不存在'}), 404
        
        # 计算差价
        budget_cost = float(item.get('预算费用', 0) or 0)
        final_cost = float(item.get('最终花费', 0) or 0)
        item['差价'] = str(budget_cost - final_cost)
        
        # 获取分类名
        category_name = item.get('category', existing_item.get('category_name'))
        
        # 更新项目
        update_item(item_id, item, category_name)
        
        return jsonify({'success': True, 'message': '更新成功'})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/delete', methods=['POST'])
def delete_item_route():
    """删除项目"""
    try:
        data = request.json
        item_ids = data.get('item_ids', [])
        
        if not item_ids:
            return jsonify({'error': '请选择要删除的项目'}), 400
        
        # 删除项目（会自动保护分类行和合计行）
        result_message = delete_items(item_ids)
        
        if result_message:
            return jsonify({'success': True, 'message': result_message})
        else:
            return jsonify({'success': True, 'message': '删除成功'})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/import', methods=['POST'])
def import_file():
    """导入Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': '不支持的文件格式，请上传 .xlsx 或 .xls 文件'}), 400
        
        # 保存上传的文件
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        upload_filename = f'import_{timestamp}_{filename}'
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], upload_filename)
        file.save(upload_path)
        
        # 验证文件格式
        validation = validate_excel_format(upload_path)
        
        if not validation['valid']:
            # 删除无效文件
            os.remove(upload_path)
            return jsonify({
                'success': False,
                'error': '文件格式验证失败',
                'errors': validation['errors'],
                'warnings': validation['warnings']
            }), 400
        
        # 解析Excel文件并导入到数据库
        # 临时使用上传的文件路径进行解析
        global EXCEL_FILE
        original_excel_file = EXCEL_FILE
        try:
            # 临时设置Excel文件路径为上传的文件
            EXCEL_FILE = upload_path
            
            # 解析Excel数据
            excel_data = parse_excel()
            
            # 导入到数据库
            import_from_excel_data(excel_data)
            
            # 恢复原Excel文件路径
            EXCEL_FILE = original_excel_file
        except Exception as e:
            EXCEL_FILE = original_excel_file
            raise e
        
        # 删除临时上传文件
        os.remove(upload_path)
        
        # 从数据库获取统计信息
        data = get_data_for_api()
        
        return jsonify({
            'success': True,
            'message': '导入成功',
            'category_count': len(data['categories']),
            'item_count': len(data['items']),
            'warnings': validation['warnings']
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': f'导入失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/validate', methods=['POST'])
def validate_file():
    """验证Excel文件格式（不导入）"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': '未选择文件'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': '不支持的文件格式，请上传 .xlsx 或 .xls 文件'}), 400
        
        # 保存临时文件
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        upload_filename = f'validate_{timestamp}_{filename}'
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], upload_filename)
        file.save(upload_path)
        
        # 验证文件格式
        validation = validate_excel_format(upload_path)
        
        # 删除临时文件
        os.remove(upload_path)
        
        return jsonify({
            'success': True,
            'valid': validation['valid'],
            'errors': validation['errors'],
            'warnings': validation['warnings'],
            'category_count': validation.get('category_count', 0),
            'has_data': validation.get('has_data', False)
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': f'验证失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500

def add_grand_total_to_excel(file_path):
    """在Excel文件末尾添加总合计行"""
    # 先更新所有合计行，确保合计值是最新的
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 找到所有分类行和对应的合计行，直接累加数据行的值
    grand_total_1st = 0
    grand_total_2nd = 0
    grand_total_actual = 0
    grand_total_diff = 0
    
    header_row = 3  # 表头行（第4行，索引3）
    
    # 直接遍历所有数据行，累加所有数字序号行的费用
    for i in range(header_row + 1, ws.max_row + 1):
        first_cell = safe_get_cell_value(ws, i, 1)
        # 检查是否是数据行（序号是数字）
        if first_cell and str(first_cell).strip().isdigit():
            # 列顺序：1st预算(5), 2nd预算(6), 实际花费(7), 差价(8)
            val_1st = safe_get_cell_value(ws, i, 5)
            val_2nd = safe_get_cell_value(ws, i, 6)
            val_actual = safe_get_cell_value(ws, i, 7)
            
            if val_1st and isinstance(val_1st, (int, float)):
                grand_total_1st += float(val_1st)
            if val_2nd and isinstance(val_2nd, (int, float)):
                grand_total_2nd += float(val_2nd)
            if val_actual and isinstance(val_actual, (int, float)):
                grand_total_actual += float(val_actual)
    
    # 计算总差价（2nd预算 - 实际花费）
    grand_total_diff = grand_total_2nd - grand_total_actual
    
    # 在文件末尾添加总合计行
    insert_row = ws.max_row + 1
    
    # 添加空行分隔
    ws.insert_rows(insert_row)
    insert_row += 1
    
    # 添加总合计行
    ws.cell(insert_row, 1, value='总计')
    safe_set_cell_value(ws, insert_row, 5, grand_total_1st if grand_total_1st > 0 else None)  # 1st预算
    safe_set_cell_value(ws, insert_row, 6, grand_total_2nd if grand_total_2nd > 0 else None)  # 2nd预算
    safe_set_cell_value(ws, insert_row, 7, grand_total_actual if grand_total_actual > 0 else None)  # 实际花费
    safe_set_cell_value(ws, insert_row, 8, grand_total_diff if grand_total_diff != 0 else None)  # 差价
    
    wb.save(file_path)
    wb.close()

def rebuild_excel_from_data():
    """基于数据库数据重新构建Excel文件"""
    # 从数据库获取当前数据
    data = get_data_for_api()
    categories = data['categories']
    items = data['items']
    
    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 40
    
    # 先计算总合计
    grand_total_1st = 0
    grand_total_2nd = 0
    grand_total_actual = 0
    grand_total_diff = 0
    
    # 累加所有项目的值
    for item in items:
        val_1st = float(item.get('1st预算费用', 0) or 0) if item.get('1st预算费用') and str(item.get('1st预算费用')).replace('.','').isdigit() else 0
        val_2nd = float(item.get('2nd预算费用', 0) or 0) if item.get('2nd预算费用') and str(item.get('2nd预算费用')).replace('.','').isdigit() else val_1st
        val_actual = float(item.get('最终实际花费', 0) or 0) if item.get('最终实际花费') and str(item.get('最终实际花费')).replace('.','').isdigit() else 0
        
        grand_total_1st += val_1st
        grand_total_2nd += val_2nd
        grand_total_actual += val_actual
    
    grand_total_diff = grand_total_2nd - grand_total_actual
    
    # 在文件开头添加总合计行
    current_row = 1
    
    # 添加总合计行
    ws.cell(current_row, 1, value='总计')
    ws.cell(current_row, 5, value=grand_total_1st if grand_total_1st > 0 else None)
    ws.cell(current_row, 6, value=grand_total_2nd if grand_total_2nd > 0 else None)
    ws.cell(current_row, 7, value=grand_total_actual if grand_total_actual > 0 else None)
    ws.cell(current_row, 8, value=grand_total_diff if grand_total_diff != 0 else None)
    current_row += 1
    
    # 添加空行分隔
    current_row += 1
    
    # 按分类组织数据
    items_by_category = {}
    for item in items:
        cat = item.get('category', '未分类')
        if cat not in items_by_category:
            items_by_category[cat] = []
        items_by_category[cat].append(item)
    
    # 数字到中文前缀的映射
    num_to_prefix = {1: '一、', 2: '二、', 3: '三、', 4: '四、', 5: '五、', 6: '六、', 7: '七、', 8: '八、', 9: '九、', 10: '十、'}
    
    # 遍历所有分类（包括前端显示的）
    for idx, category in enumerate(categories, 1):
        prefix = num_to_prefix.get(idx, '')
        
        # 添加分类行
        ws.cell(current_row, 1, value=f"{prefix}{category}")
        current_row += 1
        
        # 添加表头行
        ws.cell(current_row, 1, value='序号')
        ws.cell(current_row, 2, value='项目')
        ws.cell(current_row, 3, value='单位')
        ws.cell(current_row, 4, value='预算数量')
        ws.cell(current_row, 5, value='预算费用')
        ws.cell(current_row, 6, value='当前投入')
        ws.cell(current_row, 7, value='最终花费')
        ws.cell(current_row, 8, value='差价')
        ws.cell(current_row, 9, value='备注：选购意向（网购/实体店，品牌，型号等）')
        current_row += 1
        
        # 添加该分类下的项目
        category_items = items_by_category.get(category, [])
        category_total_budget = 0
        category_total_current = 0
        category_total_final = 0
        category_total_diff = 0
        
        # 每个分类的序号从1开始重新生成
        seq_num_in_category = 0
        
        for item in category_items:
            # 解析数值（兼容旧格式）
            val_1st = float(item.get('1st预算费用', 0) or 0) if item.get('1st预算费用') and str(item.get('1st预算费用')).replace('.','').replace('-','').isdigit() else 0
            val_2nd = float(item.get('2nd预算费用', 0) or 0) if item.get('2nd预算费用') and str(item.get('2nd预算费用')).replace('.','').replace('-','').isdigit() else 0
            val_budget = float(item.get('预算费用', 0) or 0) if item.get('预算费用') and str(item.get('预算费用')).replace('.','').replace('-','').isdigit() else (val_2nd if val_2nd > 0 else val_1st)
            val_current = float(item.get('当前投入', 0) or 0) if item.get('当前投入') and str(item.get('当前投入')).replace('.','').replace('-','').isdigit() else (float(item.get('最终实际花费', 0) or 0) if item.get('最终实际花费') and str(item.get('最终实际花费')).replace('.','').replace('-','').isdigit() else 0)
            val_final = float(item.get('最终花费', 0) or 0) if item.get('最终花费') and str(item.get('最终花费')).replace('.','').replace('-','').isdigit() else 0
            
            # 计算差价（预算费用 - 最终花费）
            val_diff = val_budget - val_final
            
            # 累加分类合计
            category_total_budget += val_budget
            category_total_current += val_current
            category_total_final += val_final
            category_total_diff += val_diff
            
            # 重新生成序号（每个分类从1开始）
            seq_num_in_category += 1
            
            ws.cell(current_row, 1, value=seq_num_in_category)
            ws.cell(current_row, 2, value=item.get('项目', ''))
            ws.cell(current_row, 3, value=item.get('单位', '') if item.get('单位') else None)
            ws.cell(current_row, 4, value=item.get('预算数量', '') if item.get('预算数量') else None)
            ws.cell(current_row, 5, value=val_budget if val_budget > 0 else None)
            ws.cell(current_row, 6, value=val_current if val_current > 0 else None)
            ws.cell(current_row, 7, value=val_final if val_final > 0 else None)
            ws.cell(current_row, 8, value=val_diff if val_diff != 0 else None)
            ws.cell(current_row, 9, value=item.get('备注', '') if item.get('备注') else None)
            current_row += 1
        
        # 添加分类合计行
        ws.cell(current_row, 1, value='合计')
        ws.cell(current_row, 5, value=category_total_budget if category_total_budget > 0 else None)
        ws.cell(current_row, 6, value=category_total_current if category_total_current > 0 else None)
        ws.cell(current_row, 7, value=category_total_final if category_total_final > 0 else None)
        ws.cell(current_row, 8, value=category_total_diff if category_total_diff != 0 else None)
        current_row += 1
    
    return wb

def register_chinese_fonts():
    """注册中文字体（只执行一次，使用缓存）"""
    global _CHINESE_FONT_REGISTERED, _CHINESE_FONT_NAME, _TEMP_FONT_FILE
    
    # 如果已经注册过，直接返回缓存的结果
    if _CHINESE_FONT_REGISTERED:
        return _CHINESE_FONT_REGISTERED, _TEMP_FONT_FILE
    
    import warnings
    import logging
    
    # 抑制fonttools的警告信息
    logging.getLogger('fontTools').setLevel(logging.ERROR)
    warnings.filterwarnings('ignore', category=UserWarning)
    
    # 尝试注册中文字体
    font_registered = False
    temp_font_file = None
    
    # 优先使用项目本地的字体文件（最快，无需提取）
    # 优先使用PingFang（更现代、更美观）
    local_font_paths = [
        os.path.join(os.path.dirname(__file__), 'fonts', 'PingFang-Regular.ttf'),  # 苹方（优先，更现代美观）
        os.path.join(os.path.dirname(__file__), 'fonts', 'SimHei.ttf'),  # 黑体（已提取的PingFang）
        os.path.join(os.path.dirname(__file__), 'fonts', 'SimSun.ttf'),   # 宋体
        os.path.join(os.path.dirname(__file__), 'fonts', 'Arial Unicode.ttf'),  # Arial Unicode（支持中文）
        os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSansCJK-Regular.ttf'),  # 思源黑体
        os.path.join(os.path.dirname(__file__), 'fonts', 'SourceHanSansCN-Regular.otf'),  # 思源黑体OTF
    ]
    
    for font_path in local_font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                font_registered = True
                break
            except Exception as e:
                continue
    
    # 如果本地字体不可用，尝试系统字体
    if not font_registered:
        import platform
        import tempfile
        
        if platform.system() == 'Darwin':  # macOS
            # 优先尝试TTF文件
            ttf_paths = [
                '/Library/Fonts/Microsoft/SimHei.ttf',  # 黑体
                '/Library/Fonts/Microsoft/SimSun.ttf',  # 宋体
            ]
            
            for font_path in ttf_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                        font_registered = True
                        break
                    except:
                        continue
            
            # 如果TTF文件不可用，尝试从TTC提取（最慢）
            if not font_registered:
                ttc_paths = [
                    '/System/Library/Fonts/PingFang.ttc',  # 苹方
                    '/System/Library/Fonts/STHeiti Light.ttc',  # 黑体
                    '/System/Library/Fonts/STHeiti Medium.ttc',
                ]
                
                try:
                    from fontTools.ttLib import TTFont as FTTTFont
                    
                    # 抑制fonttools的警告
                    import sys
                    original_stderr = sys.stderr
                    try:
                        # 临时重定向stderr以抑制警告
                        sys.stderr = open(os.devnull, 'w')
                        
                        for ttc_path in ttc_paths:
                            if os.path.exists(ttc_path):
                                try:
                                    # 从TTC文件中提取第一个字体（编号0）
                                    ttc = FTTTFont(ttc_path, fontNumber=0)
                                    if len(ttc.getGlyphSet()) > 0:
                                        # 创建临时TTF文件
                                        temp_font_file = tempfile.NamedTemporaryFile(delete=False, suffix='.ttf')
                                        ttc.save(temp_font_file.name)
                                        temp_font_file.close()
                                        
                                        # 注册临时字体文件
                                        pdfmetrics.registerFont(TTFont('ChineseFont', temp_font_file.name))
                                        font_registered = True
                                        break
                                except Exception as e:
                                    continue
                    finally:
                        # 恢复stderr
                        sys.stderr.close()
                        sys.stderr = original_stderr
                except ImportError:
                    # fonttools未安装，跳过TTC处理
                    pass
    
    # 缓存结果
    _CHINESE_FONT_REGISTERED = font_registered
    _CHINESE_FONT_NAME = 'ChineseFont' if font_registered else 'Helvetica'
    _TEMP_FONT_FILE = temp_font_file
    
    return font_registered, temp_font_file

def generate_pdf():
    """使用reportlab生成PDF（从数据库读取数据）"""
    global _CHINESE_FONT_NAME
    
    import warnings
    import logging
    
    # 抑制fonttools和reportlab的警告
    warnings.filterwarnings('ignore', category=UserWarning)
    warnings.filterwarnings('ignore', message='.*CFF.*')
    logging.getLogger('fontTools').setLevel(logging.ERROR)
    
    # 注册中文字体（使用缓存，只注册一次）
    chinese_font_registered, temp_font_file = register_chinese_fonts()
    chinese_font_name = _CHINESE_FONT_NAME
    
    try:
        # 从数据库获取数据
        data = get_data_for_api()
        categories = data['categories']
        items = data['items']
        
        # 按分类组织数据
        items_by_category = {}
        for item in items:
            cat = item.get('category', '未分类')
            if cat not in items_by_category:
                items_by_category[cat] = []
            items_by_category[cat].append(item)
        
        # 计算总合计（使用新字段名）
        grand_total_budget = 0
        grand_total_current = 0
        grand_total_final = 0
        grand_total_diff = 0
        
        for item in items:
            val_budget = float(item.get('预算费用', 0) or 0) if item.get('预算费用') and str(item.get('预算费用')).replace('.','').isdigit() else 0
            val_current = float(item.get('当前投入', 0) or 0) if item.get('当前投入') and str(item.get('当前投入')).replace('.','').isdigit() else 0
            val_final = float(item.get('最终花费', 0) or 0) if item.get('最终花费') and str(item.get('最终花费')).replace('.','').isdigit() else 0
            
            grand_total_budget += val_budget
            grand_total_current += val_current
            grand_total_final += val_final
        
        grand_total_diff = grand_total_budget - grand_total_final
        
        def format_number(value):
            """格式化数字"""
            if not value or str(value).strip() == '':
                return '0.00'
            try:
                num = float(value)
                return f'{num:,.2f}'
            except:
                return '0.00'
        
        # 创建PDF文档
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_filename = f'红玺台复式装修预算表_导出_{timestamp}.pdf'
        export_path = os.path.join(app.config['EXPORT_FOLDER'], export_filename)
        
        doc = SimpleDocTemplate(export_path, pagesize=A4,
                               rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        
        # 创建样式（使用中文字体）
        styles = getSampleStyleSheet()
        
        # 自定义样式（使用中文字体）
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=chinese_font_name,
            fontSize=20,
            textColor=colors.HexColor('#667eea'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontName=chinese_font_name,
            fontSize=14,
            textColor=colors.white,
            backColor=colors.HexColor('#667eea'),
            alignment=TA_LEFT,
            spaceAfter=10,
            spaceBefore=10,
            leftIndent=10,
            rightIndent=10
        )
        
        summary_style = ParagraphStyle(
            'CustomSummary',
            parent=styles['Normal'],
            fontName=chinese_font_name,
            fontSize=11,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=15
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=chinese_font_name,
            fontSize=9
        )
        
        # 备注样式（更小的字体）
        remark_style = ParagraphStyle(
            'Remark',
            parent=styles['Normal'],
            fontName=chinese_font_name,
            fontSize=7,  # 备注使用7号字体，比正文小
            textColor=colors.HexColor('#666666'),
            leading=8  # 行距
        )
        
        time_style = ParagraphStyle(
            'Time',
            parent=styles['Normal'],
            fontName=chinese_font_name,
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontName=chinese_font_name,
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        
        def truncate_text(text, max_length=25):
            """截断文本，超出部分用...显示"""
            if not text:
                return ''
            text = str(text).strip()
            if len(text) <= max_length:
                return text
            return text[:max_length-3] + '...'
        
        # 构建PDF内容
        story = []
        
        # 标题
        story.append(Paragraph('装修预算表', title_style))
        story.append(Paragraph(f'生成时间：{datetime.now().strftime("%Y年%m月%d日 %H:%M")}', time_style))
        story.append(Spacer(1, 0.5*cm))
        
        # 总合计（使用reportlab支持的颜色格式）
        diff_color = "green" if grand_total_diff >= 0 else "red"
        total_text = f'<b>总合计：</b> 预算费用 <b><font color=blue>{format_number(grand_total_budget)}</font></b> 元 | ' \
                     f'当前投入 <b><font color=orange>{format_number(grand_total_current)}</font></b> 元 | ' \
                     f'最终花费 <b><font color=green>{format_number(grand_total_final)}</font></b> 元 | ' \
                     f'差价 <b><font color={diff_color}>{format_number(grand_total_diff)}</font></b> 元'
        
        total_table = Table([[Paragraph(total_text, summary_style)]], 
                            colWidths=[16*cm],
                            style=TableStyle([
                                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
                                ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#667eea')),
                                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                                ('TOPPADDING', (0, 0), (-1, -1), 15),
                                ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
                            ]))
        story.append(total_table)
        story.append(Spacer(1, 0.5*cm))
        
        # 遍历所有分类
        for category in categories:
            category_items = items_by_category.get(category, [])
            category_total_budget = 0
            category_total_current = 0
            category_total_final = 0
            category_total_diff = 0
            
            # 分类标题
            category_header = Table([[Paragraph(f'{category} <font size=10>({len(category_items)} 项)</font>', header_style)]],
                                    colWidths=[16*cm],
                                    style=TableStyle([
                                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#667eea')),
                                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                        ('LEFTPADDING', (0, 0), (-1, -1), 10),
                                        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                                        ('TOPPADDING', (0, 0), (-1, -1), 10),
                                        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                                    ]))
            story.append(category_header)
            
            if len(category_items) > 0:
                # 表头（使用中文字体）
                table_data = [['序号', '项目名称', '单位', '数量', '预算费用', '当前投入', '最终花费', '差价', '备注']]
                
                seq_num = 0
                for item in category_items:
                    seq_num += 1
                    # 兼容旧格式
                    val_1st = float(item.get('1st预算费用', 0) or 0) if item.get('1st预算费用') and str(item.get('1st预算费用')).replace('.','').replace('-','').isdigit() else 0
                    val_2nd = float(item.get('2nd预算费用', 0) or 0) if item.get('2nd预算费用') and str(item.get('2nd预算费用')).replace('.','').replace('-','').isdigit() else 0
                    val_budget = float(item.get('预算费用', 0) or 0) if item.get('预算费用') and str(item.get('预算费用')).replace('.','').replace('-','').isdigit() else (val_2nd if val_2nd > 0 else val_1st)
                    val_current = float(item.get('当前投入', 0) or 0) if item.get('当前投入') and str(item.get('当前投入')).replace('.','').replace('-','').isdigit() else (float(item.get('最终实际花费', 0) or 0) if item.get('最终实际花费') and str(item.get('最终实际花费')).replace('.','').replace('-','').isdigit() else 0)
                    val_final = float(item.get('最终花费', 0) or 0) if item.get('最终花费') and str(item.get('最终花费')).replace('.','').replace('-','').isdigit() else 0
                    val_diff = val_budget - val_final
                    
                    category_total_budget += val_budget
                    category_total_current += val_current
                    category_total_final += val_final
                    category_total_diff += val_diff
                    
                    # 处理备注：截断并用小字体显示
                    remark_text = truncate_text(item.get('备注', ''), max_length=25)
                    remark_cell = Paragraph(remark_text, remark_style) if remark_text else ''
                    
                    table_data.append([
                        str(seq_num),
                        Paragraph(item.get('项目', ''), normal_style),
                        item.get('单位', ''),
                        item.get('预算数量', ''),
                        format_number(val_budget),
                        format_number(val_current),
                        format_number(val_final),
                        format_number(val_diff),
                        remark_cell  # 使用Paragraph样式，字体更小
                    ])
                
                # 创建表格
                table = Table(table_data, colWidths=[0.8*cm, 4*cm, 1*cm, 1*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2*cm],
                             repeatRows=1)
                
                table.setStyle(TableStyle([
                    # 表头样式
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (4, 0), (7, -1), 'RIGHT'),  # 数字列右对齐
                    ('FONTNAME', (0, 0), (-1, 0), chinese_font_name),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    # 边框
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
                    # 数据行样式（除了备注列）
                    ('FONTNAME', (0, 1), (-1, -1), chinese_font_name),
                    ('FONTSIZE', (0, 1), (7, -1), 8),  # 前8列（不包括备注）使用8号字体
                    ('FONTSIZE', (8, 1), (8, -1), 7),  # 备注列（第9列）使用7号字体
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                
                story.append(table)
            
            # 分类合计
            summary_text = f'本分类合计：预算费用 <b>{format_number(category_total_budget)}</b> 元 | ' \
                          f'当前投入 <b>{format_number(category_total_current)}</b> 元 | ' \
                          f'最终花费 <b>{format_number(category_total_final)}</b> 元 | ' \
                          f'差价 <b>{format_number(category_total_diff)}</b> 元'
            
            summary_table = Table([[Paragraph(summary_text, summary_style)]],
                                 colWidths=[16*cm],
                                 style=TableStyle([
                                     ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
                                     ('LINEABOVE', (0, 0), (-1, -1), 2, colors.grey),
                                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                     ('LEFTPADDING', (0, 0), (-1, -1), 10),
                                     ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                                     ('TOPPADDING', (0, 0), (-1, -1), 8),
                                     ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                                 ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3*cm))
        
        # 页脚
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph('本预算表由装修预算表管理系统自动生成', footer_style))
        
        # 构建PDF
        doc.build(story)
        
        # 清理临时字体文件
        if temp_font_file and os.path.exists(temp_font_file.name):
            try:
                os.unlink(temp_font_file.name)
            except:
                pass
        
        return export_path
        
    except Exception as e:
        # 清理临时字体文件
        if temp_font_file and os.path.exists(temp_font_file.name):
            try:
                os.unlink(temp_font_file.name)
            except:
                pass
        raise

@app.route('/api/export', methods=['GET'])
def export_file():
    """导出Excel文件（基于前端数据重新构建）"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_filename = f'红玺台复式装修预算表_导出_{timestamp}.xlsx'
        export_path = os.path.join(app.config['EXPORT_FOLDER'], export_filename)
        
        # 基于前端数据重新构建Excel
        wb = rebuild_excel_from_data()
        wb.save(export_path)
        wb.close()
        
        return send_file(
            export_path,
            as_attachment=True,
            download_name=export_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/export-pdf', methods=['GET'])
def export_pdf():
    """导出PDF文件（基于前端数据生成）"""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'PDF导出功能不可用，请安装reportlab: pip install reportlab'}), 500
    
    try:
        # 生成PDF
        export_path = generate_pdf()
        export_filename = os.path.basename(export_path)
        
        return send_file(
            export_path,
            as_attachment=True,
            download_name=export_filename,
            mimetype='application/pdf'
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

def parse_text_local(text):
    """使用本地规则解析自然语言文本，提取装修项目信息"""
    try:
        # 从数据库获取现有分类
        data = get_data_for_api()
        categories = data['categories']
        
        result = {
            '项目': '',
            'category': '',
            '单位': '',
            '预算数量': '',
            '预算费用': '',  # 使用新字段名
            '当前投入': '',  # 使用新字段名
            '最终花费': '',  # 使用新字段名
            '备注': ''
        }
        
        # 提取分类名和项目名称
        # 如果文本包含逗号，第一个部分用于分类匹配，第二个部分用于项目名
        parts = re.split(r'[，,，]', text)
        parts = [p.strip() for p in parts if p.strip()]  # 去除空部分
        
        if len(parts) >= 2:
            # 第一个部分用于分类匹配
            category_candidate = parts[0]
            # 从第二个部分开始查找项目名（跳过数量和预算信息）
            project_name = None
            for i in range(1, len(parts)):
                part = parts[i].strip()
                # 如果这个部分看起来像数量（数字+单位），跳过
                if re.match(r'^\d+(?:\.\d+)?\s*(?:套|个|米|平方米|平方厘米|件|台|张|把|支|根|条|块|片|组|项)$', part):
                    continue
                # 如果这个部分看起来像预算/费用信息，跳过
                if re.match(r'^(?:预算|当前|实际|最终|花费|费用|投入)[：:，,，\s]*\d+', part):
                    continue
                # 否则作为项目名
                project_name = part
                break
            
            # 如果没找到合适的项目名，使用第一个部分作为项目名（第一个部分可能是项目名而不是分类名）
            if not project_name:
                # 检查第一个部分是否看起来像分类名（包含"全屋"、"定制"等关键词）
                if any(kw in category_candidate for kw in ['全屋', '定制', '基装', '智能', '家居']):
                    # 第一个部分像分类名，但没有找到项目名，使用第二个部分（即使它可能是数量）
                    if len(parts) >= 2:
                        project_name = parts[1]
                else:
                    # 第一个部分不像分类名，可能是项目名
                    project_name = category_candidate
                    category_candidate = None  # 清空分类候选，后续根据项目名匹配
            
            if project_name:
                result['项目'] = project_name
            # 先尝试将第一个部分作为分类候选（后续会进行匹配）
            if category_candidate:
                result['category'] = category_candidate
        else:
            # 如果只有一个部分，需要智能提取项目名（排除数字、单位、预算等关键词）
            # 提取项目名：从开头到第一个数字或关键词之前
            # 先尝试匹配到第一个数字或关键词之前的内容
            project_match = re.search(r'^([^，,，\d预算当前实际最终备注]+?)(?=[，,，]|\d|预算|当前|实际|最终|备注|$)', text)
            if project_match:
                project_name = project_match.group(1).strip()
                # 清理项目名（去除末尾的单位、数字等）
                project_name = re.sub(r'\s*\d+(?:\.\d+)?\s*(?:套|个|米|平方米|平方厘米|件|台|张|把|支|根|条|块|片|组|项)?\s*$', '', project_name)
                if project_name:
                    result['项目'] = project_name
            else:
                # 如果正则匹配失败，尝试直接提取第一个非数字、非关键词的部分
                # 找到第一个逗号、数字或关键词的位置
                match = re.search(r'^([^，,，\d预算当前实际最终备注]+)', text)
                if match:
                    project_name = match.group(1).strip()
                    # 清理项目名
                    project_name = re.sub(r'\s*\d+(?:\.\d+)?\s*(?:套|个|米|平方米|平方厘米|件|台|张|把|支|根|条|块|片|组|项)?\s*$', '', project_name)
                    if project_name:
                        result['项目'] = project_name
        
        # 提取数量（数字+单位，如"1套"、"3个"）
        quantity_match = re.search(r'(\d+(?:\.\d+)?)\s*([套个米平方米平方厘米件台张把支根条块片组项])', text)
        if quantity_match:
            result['预算数量'] = quantity_match.group(1)
            result['单位'] = quantity_match.group(2)
        
        # 提取费用（数字+元，或单独的数字）
        # 预算费用（优先匹配"预算"关键词，支持"预算59000元"格式）
        budget_match = re.search(r'预算\s*(\d+(?:\.\d+)?)\s*元?', text)
        if budget_match:
            result['预算费用'] = budget_match.group(1)
        else:
            # 如果没有明确标识，尝试找第一个费用数字（在"元"之前）
            cost_match = re.search(r'(\d+(?:\.\d+)?)\s*元', text)
            if cost_match:
                result['预算费用'] = cost_match.group(1)
        
        # 当前投入（匹配"当前投入"、"当前花费"等，支持"当前投入2000元"格式）
        current_match = re.search(r'当前投入\s*(\d+(?:\.\d+)?)\s*元?', text)
        if current_match:
            result['当前投入'] = current_match.group(1)
        else:
            result['当前投入'] = '0'
        
        # 最终花费（匹配"最终花费"、"实际花费"、"实际费用"等）
        final_match = re.search(r'(?:最终花费|最终费用|实际花费|实际费用)\s*(\d+(?:\.\d+)?)\s*元?', text)
        if final_match:
            result['最终花费'] = final_match.group(1)
        else:
            result['最终花费'] = '0'
        
        # 提取备注（在"备注"、"品牌"、"型号"等关键词之后）
        remark_keywords = ['备注', '品牌', '型号', '渠道', '介绍', '说明']
        for keyword in remark_keywords:
            remark_match = re.search(rf'{keyword}[：:，,，\s]+([^，,，]+)', text)
            if remark_match:
                remark_text = remark_match.group(1).strip()
                # 排除纯数字（可能是误匹配）
                if not re.match(r'^\d+(?:\.\d+)?\s*(?:元)?$', remark_text):
                    result['备注'] = remark_text
                    break
        
        # 如果没有找到备注，尝试提取最后一部分作为备注（排除数字和"实际XX"）
        if not result['备注']:
            # 提取最后一个逗号后的内容
            parts = re.split(r'[，,，]', text)
            if len(parts) > 1:
                last_part = parts[-1].strip()
                # 排除纯数字、"实际XX"这样的内容
                if (not re.match(r'^\d+(?:\.\d+)?\s*(?:元|套|个|米)?$', last_part) and 
                    not re.match(r'^实际\s*\d+', last_part)):
                    result['备注'] = last_part
        
        # 自动匹配分类（根据项目名称关键词）
        if not result['category'] and result['项目']:
            project_name = result['项目']
            matched = False
            # 先尝试精确匹配分类名（如果项目名完全匹配某个分类）
            for cat in categories:
                if project_name == cat or cat in project_name:
                    result['category'] = cat
                    matched = True
                    break
            
            # 如果没精确匹配，尝试模糊匹配（关键词匹配 + 相似度匹配）
            if not matched:
                best_match = None
                best_score = 0
                
                for cat in categories:
                    score = 0
                    
                    # 1. 关键词匹配（提高权重）
                    # 基装相关
                    if any(kw in project_name for kw in ['基装', '基础', '装修', '吊顶', '改水', '改电', '土建', '乳胶漆', '防锈漆', '楼梯']):
                        if any(kw in cat for kw in ['基装', '基础', '装修']):
                            score = max(score, 0.8)
                    # 柜/定制相关（支持部分匹配）
                    if any(kw in project_name for kw in ['柜', '衣柜', '鞋柜', '橱柜', '定制']):
                        if any(kw in cat for kw in ['柜', '定制']):
                            score = max(score, 0.8)
                    # 电器相关
                    if any(kw in project_name for kw in ['电器', '家电', '智能', '空调', '冰箱', '洗衣机', '电视', '家居']):
                        if any(kw in cat for kw in ['电', '智能', '电器', '家居']):
                            score = max(score, 0.8)
                    # 卫浴相关
                    if any(kw in project_name for kw in ['卫浴', '浴室', '卫生间', '马桶', '花洒', '洗手盆', '淋浴']):
                        if any(kw in cat for kw in ['卫浴', '浴室', '卫生间']):
                            score = max(score, 0.8)
                    # 地板相关
                    if any(kw in project_name for kw in ['地板', '地砖', '瓷砖', '木地板']):
                        if any(kw in cat for kw in ['地板', '地砖', '瓷砖']):
                            score = max(score, 0.8)
                    # 门窗相关
                    if any(kw in project_name for kw in ['门', '窗', '门窗', '防盗门', '铝合金']):
                        if any(kw in cat for kw in ['门', '窗', '门窗']):
                            score = max(score, 0.8)
                    
                    # 2. 字符相似度匹配（作为补充）
                    # 提取关键词（去除常见修饰词）
                    project_chars = set([c for c in project_name if c not in ['的', '和', '与', '及', '或', '、', ',', '，', ' ', '全', '屋']])
                    cat_chars = set([c for c in cat if c not in ['的', '和', '与', '及', '或', '、', ',', '，', ' ', '全', '屋']])
                    common_chars = project_chars & cat_chars
                    if common_chars:
                        max_len = max(len(project_chars), len(cat_chars))
                        if max_len > 0:
                            char_similarity = len(common_chars) / max_len
                            # 如果字符相似度较高，提高分数
                            if char_similarity > 0.3:
                                score = max(score, char_similarity * 0.6)  # 字符相似度权重较低
                    
                    # 3. 部分匹配（如果项目名包含分类名的关键词，或分类名包含项目名的关键词）
                    # 提取项目名和分类名的核心词（2个字符以上）
                    project_words = [project_name[i:i+2] for i in range(len(project_name)-1)]
                    cat_words = [cat[i:i+2] for i in range(len(cat)-1)]
                    common_words = set(project_words) & set(cat_words)
                    if common_words:
                        word_similarity = len(common_words) / max(len(set(project_words)), len(set(cat_words)))
                        if word_similarity > 0.2:
                            score = max(score, word_similarity * 0.7)
                    
                    # 记录最佳匹配
                    if score > best_score:
                        best_score = score
                        best_match = cat
                
                # 如果找到最佳匹配且相似度超过阈值，使用它
                if best_match and best_score >= 0.3:
                    result['category'] = best_match
                    matched = True
            
            # 如果还没匹配到，使用项目名称作为新分类名称
            if not matched:
                # 提取项目名称的关键部分作为分类名（去掉数量、单位等）
                category_name = result['项目']
                # 如果项目名包含逗号，取第一部分
                if '，' in category_name or ',' in category_name:
                    category_name = category_name.split('，')[0].split(',')[0].strip()
                # 如果项目名太长，截取前10个字符
                if len(category_name) > 10:
                    category_name = category_name[:10]
                result['category'] = category_name
        
        # 清理空值
        for key in result:
            if not result[key]:
                result[key] = ''
        
        return {'success': True, 'item': result}
        
    except Exception as e:
        import traceback
        return {'error': f'解析失败: {str(e)}', 'traceback': traceback.format_exc()}

@app.route('/api/parse', methods=['POST'])
def parse_text():
    """本地解析自然语言输入（支持批量）"""
    try:
        data = request.json
        text = data.get('text', '').strip()
        
        if not text:
            return jsonify({'error': '请输入文本'}), 400
        
        # 支持批量解析：按换行符分割
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        if len(lines) == 1:
            # 单个项目
            result = parse_text_local(lines[0])
            if 'error' in result:
                return jsonify(result), 500
            return jsonify({'success': True, 'item': result.get('item'), 'items': [result.get('item')], 'is_batch': False})
        else:
            # 批量解析
            items = []
            errors = []
            for i, line in enumerate(lines, 1):
                try:
                    result = parse_text_local(line)
                    if 'item' in result:
                        items.append(result['item'])
                    else:
                        errors.append(f'第{i}行: {result.get("error", "解析失败")}')
                except Exception as e:
                    errors.append(f'第{i}行: {str(e)}')
            
            if items:
                return jsonify({
                    'success': True, 
                    'items': items, 
                    'is_batch': True,
                    'total': len(lines),
                    'success_count': len(items),
                    'error_count': len(errors),
                    'errors': errors if errors else None
                })
            else:
                return jsonify({
                    'success': False, 
                    'error': '所有项目解析失败',
                    'errors': errors
                }), 400
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/parse-and-add', methods=['POST'])
def parse_and_add():
    """智能解析并直接添加项目"""
    try:
        data = request.json
        text = data.get('text', '').strip()
        
        # 如果直接提供了item数据（从预览确认），直接使用
        if 'item' in data and data['item']:
            item = data['item']
            category = item.get('category', '')
            if 'category' in item:
                del item['category']
            
            # 计算差价
            budget_cost = float(item.get('预算费用', 0) or 0)
            final_cost = float(item.get('最终花费', 0) or 0)
            item['差价'] = str(budget_cost - final_cost)
            
            # 调用数据库模块的add_item函数
            from database import add_item as db_add_item
            db_add_item(item, category)
            return jsonify({
                'success': True,
                'message': '智能添加成功',
                'item': item,
                'category': category
            })
        
        if not text:
            return jsonify({'error': '请输入文本'}), 400
        
        # 支持批量解析：按换行符分割
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        if len(lines) == 1:
            # 单个项目
            parse_result = parse_text_local(lines[0])
            if 'error' in parse_result:
                return jsonify(parse_result), 500
            
            item = parse_result['item']
            category = item.get('category', '')
            
            # 清理category字段，从item中移除
            if 'category' in item:
                del item['category']
            
            # 兼容旧字段名，转换为新字段名
            if '1st预算费用' in item and item['1st预算费用']:
                if not item.get('预算费用'):
                    item['预算费用'] = item['1st预算费用']
            if '2nd预算费用' in item and item['2nd预算费用']:
                if not item.get('预算费用'):
                    item['预算费用'] = item['2nd预算费用']
            if '最终实际花费' in item and item['最终实际花费']:
                if not item.get('最终花费'):
                    item['最终花费'] = item['最终实际花费']
            
            # 计算差价
            budget_cost = float(item.get('预算费用', 0) or 0)
            final_cost = float(item.get('最终花费', 0) or 0)
            item['差价'] = str(budget_cost - final_cost)
            
            # 添加到数据库
            from database import add_item as db_add_item
            db_add_item(item, category)
            
            return jsonify({
                'success': True,
                'message': '智能添加成功',
                'item': item,
                'category': category,
                'count': 1
            })
        else:
            # 批量添加
            from database import add_item as db_add_item
            success_count = 0
            error_count = 0
            errors = []
            
            for i, line in enumerate(lines, 1):
                try:
                    parse_result = parse_text_local(line)
                    if 'error' in parse_result:
                        error_count += 1
                        errors.append(f'第{i}行: {parse_result.get("error", "解析失败")}')
                        continue
                    
                    item = parse_result['item']
                    category = item.get('category', '')
                    
                    # 清理category字段
                    if 'category' in item:
                        del item['category']
                    
                    # 兼容旧字段名
                    if '1st预算费用' in item and item['1st预算费用']:
                        if not item.get('预算费用'):
                            item['预算费用'] = item['1st预算费用']
                    if '2nd预算费用' in item and item['2nd预算费用']:
                        if not item.get('预算费用'):
                            item['预算费用'] = item['2nd预算费用']
                    if '最终实际花费' in item and item['最终实际花费']:
                        if not item.get('最终花费'):
                            item['最终花费'] = item['最终实际花费']
                    
                    # 计算差价
                    budget_cost = float(item.get('预算费用', 0) or 0)
                    final_cost = float(item.get('最终花费', 0) or 0)
                    item['差价'] = str(budget_cost - final_cost)
                    
                    # 添加项目
                    db_add_item(item, category)
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f'第{i}行: {str(e)}')
            
            if success_count > 0:
                message = f'成功添加 {success_count} 项'
                if error_count > 0:
                    message += f'，失败 {error_count} 项'
                return jsonify({
                    'success': True,
                    'message': message,
                    'count': success_count,
                    'total': len(lines),
                    'errors': errors if errors else None
                })
            else:
                return jsonify({
                    'success': False,
                    'error': '所有项目添加失败',
                    'errors': errors
                }), 400
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

# 应用启动时预注册中文字体（避免首次PDF导出时的延迟）
def _init_fonts_on_startup():
    """在应用启动时预注册字体"""
    global _CHINESE_FONT_REGISTERED
    if REPORTLAB_AVAILABLE and not _CHINESE_FONT_REGISTERED:
        try:
            print("正在注册中文字体...", end='', flush=True)
            font_registered, _ = register_chinese_fonts()
            if font_registered:
                print(" ✓ 中文字体注册成功")
            else:
                print(" ⚠ 未找到中文字体，PDF可能无法正确显示中文")
        except Exception as e:
            print(f" ⚠ 字体注册失败: {e}")

# 如果是直接运行（不是被导入），在启动时初始化字体
if __name__ == '__main__':
    import os
    # 检查环境变量，生产环境不使用debug模式
    debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    # 默认监听本地，可通过环境变量FLASK_HOST=0.0.0.0允许局域网访问
    host = os.getenv('FLASK_HOST', '127.0.0.1')
    port = int(os.getenv('FLASK_PORT', '5000'))
    
    if REPORTLAB_AVAILABLE:
        _init_fonts_on_startup()
    app.run(debug=debug_mode, host=host, port=port)
else:
    # 如果是被导入（如gunicorn），使用异步初始化避免阻塞
    if REPORTLAB_AVAILABLE:
        import threading
        def _init_fonts_async():
            """异步初始化字体，避免阻塞应用启动"""
            _init_fonts_on_startup()
        threading.Thread(target=_init_fonts_async, daemon=True).start()
